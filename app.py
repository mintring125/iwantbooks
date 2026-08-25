import csv
import json
import os
import re
import time
import unicodedata
from datetime import datetime
from io import BytesIO, TextIOWrapper

import openpyxl
import requests
import xlrd
from flask import Flask, jsonify, render_template, request, send_file, session
from flask_sqlalchemy import SQLAlchemy
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import UniqueConstraint


def build_database_url():
    database_url = os.environ.get("DATABASE_URL", "").strip()
    if not database_url:
        return "sqlite:///school_books.db"
    if database_url.startswith("postgres://"):
        return database_url.replace("postgres://", "postgresql+psycopg://", 1)
    if database_url.startswith("postgresql://"):
        return database_url.replace("postgresql://", "postgresql+psycopg://", 1)
    return database_url


app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "purchase-wishlist-2026")
app.config["SQLALCHEMY_DATABASE_URI"] = build_database_url()
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)
db_initialized = False

ALADIN_API_KEY = os.environ.get("ALADIN_API_KEY", "").strip()
ALADIN_SEARCH_URL = os.environ.get(
    "ALADIN_SEARCH_URL",
    "https://aladin.co.kr/ttb/api/ItemSearch.aspx",
).strip()
ALADIN_LIST_URL = os.environ.get(
    "ALADIN_LIST_URL",
    "https://www.aladin.co.kr/ttb/api/ItemList.aspx",
).strip()
ALADIN_CHILDREN_CATEGORY_ID = 1108
BESTSELLER_CACHE_SECONDS = 6 * 60 * 60
BESTSELLER_RETRY_SECONDS = 10 * 60
bestseller_cache = {"expires_at": 0.0, "retry_after": 0.0, "items": []}
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "2026")
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SUBMISSIONS_FILE = os.path.join(BASE_DIR, "submissions.json")
CATALOG_FILE = os.path.join(BASE_DIR, "library_catalog.json")
EXPORT_TEMPLATE_FILE = os.path.join(
    BASE_DIR, "2026학년도 1학기 학생 및 학부모, 교직원 구입 희망도서 (   학년   반).xlsx"
)

SCHOOL_STRUCTURE = {
    "1": {"1": 11},
    "2": {"1": 15},
    "3": {"1": 11, "2": 11},
    "4": {"1": 17, "2": 16},
    "5": {"1": 17, "2": 16},
    "6": {"1": 19, "2": 20},
}
GRADE_OPTIONS = list(SCHOOL_STRUCTURE.keys())


class Submission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    grade = db.Column(db.String(2), nullable=False)
    class_num = db.Column(db.String(2), nullable=False)
    student_number = db.Column(db.String(3), nullable=False)
    student_label = db.Column(db.String(32), nullable=False)
    books_json = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    __table_args__ = (
        UniqueConstraint("grade", "class_num", "student_number", name="uq_student_slot"),
    )

    def to_dict(self):
        return {
            "grade": self.grade,
            "classNum": self.class_num,
            "studentNumber": self.student_number,
            "studentLabel": self.student_label,
            "books": json.loads(self.books_json),
            "timestamp": self.created_at.strftime("%Y-%m-%d %H:%M:%S"),
        }


class CatalogBook(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(500), nullable=False)
    normalized_title = db.Column(db.String(500), nullable=False, index=True)
    isbn = db.Column(db.String(32), nullable=True, index=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)

    def to_dict(self):
        return {"title": self.title, "isbn": self.isbn or ""}


class ApiCache(db.Model):
    key = db.Column(db.String(64), primary_key=True)
    value_json = db.Column(db.Text, nullable=False)
    updated_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


def class_options_for_grade(grade):
    return list(SCHOOL_STRUCTURE.get(str(grade), {}).keys())


def student_numbers_for_class(grade, class_num):
    max_number = SCHOOL_STRUCTURE.get(str(grade), {}).get(str(class_num), 0)
    return [str(i) for i in range(1, max_number + 1)]


def normalize_title(title):
    if not title:
        return ""
    normalized = unicodedata.normalize("NFC", str(title))
    normalized = re.sub(r"[^0-9A-Za-z가-힣]", "", normalized)
    return normalized.lower().strip()


def normalize_isbn(isbn):
    return re.sub(r"[^0-9]", "", str(isbn or ""))


def bootstrap_submissions_from_json():
    if Submission.query.count() > 0 or not os.path.exists(SUBMISSIONS_FILE):
        return

    try:
        with open(SUBMISSIONS_FILE, "r", encoding="utf-8") as file:
            raw_items = json.load(file)
    except Exception:
        return

    for item in raw_items:
        grade = str(item.get("grade", "")).strip()
        class_num = str(item.get("classNum", "")).strip()
        student_number = str(item.get("studentNumber", item.get("name", ""))).strip()
        books = item.get("books", [])
        if grade not in GRADE_OPTIONS:
            continue
        if class_num not in class_options_for_grade(grade):
            continue
        if student_number not in student_numbers_for_class(grade, class_num):
            continue

        created_at = datetime.utcnow()
        timestamp = str(item.get("timestamp", "")).strip()
        if timestamp:
            try:
                created_at = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass

        if Submission.query.filter_by(
            grade=grade, class_num=class_num, student_number=student_number
        ).first():
            continue

        db.session.add(
            Submission(
                grade=grade,
                class_num=class_num,
                student_number=student_number,
                student_label=f"{grade}학년 {class_num}반 {student_number}번",
                books_json=json.dumps(books, ensure_ascii=False),
                created_at=created_at,
            )
        )
    db.session.commit()


def bootstrap_catalog_from_json():
    if CatalogBook.query.count() > 0 or not os.path.exists(CATALOG_FILE):
        return

    try:
        with open(CATALOG_FILE, "r", encoding="utf-8") as file:
            raw_items = json.load(file)
    except Exception:
        return

    for item in raw_items:
        title = str(item.get("title", "")).strip()
        if not title:
            continue
        isbn = normalize_isbn(item.get("isbn", ""))
        db.session.add(
            CatalogBook(
                title=title,
                normalized_title=normalize_title(title),
                isbn=isbn or None,
            )
        )
    db.session.commit()


def ensure_database_ready():
    global db_initialized
    if db_initialized:
        return

    with app.app_context():
        db.create_all()
        bootstrap_submissions_from_json()
        bootstrap_catalog_from_json()
    db_initialized = True


def load_catalog_index():
    ensure_database_ready()

    rows = CatalogBook.query.with_entities(
        CatalogBook.normalized_title,
        CatalogBook.isbn,
    ).all()
    titles = {row[0] for row in rows if row[0]}
    isbns = {row[1] for row in rows if row[1]}
    return titles, isbns


def is_catalog_duplicate(book_title, book_isbn, catalog_titles, catalog_isbns):
    clean_isbn = normalize_isbn(book_isbn)
    if clean_isbn and clean_isbn in catalog_isbns:
        return True

    normalized_query = normalize_title(book_title)
    if not normalized_query:
        return False

    for normalized_catalog in catalog_titles:
        if not normalized_catalog:
            continue
        if normalized_query == normalized_catalog:
            return True
        shorter, longer = sorted(
            (normalized_query, normalized_catalog),
            key=len,
        )
        if (
            len(shorter) >= 6
            and shorter in longer
            and len(shorter) / len(longer) >= 0.65
        ):
            return True
    return False


def check_duplicate(book_title, book_isbn=""):
    catalog_titles, catalog_isbns = load_catalog_index()
    return is_catalog_duplicate(
        book_title,
        book_isbn,
        catalog_titles,
        catalog_isbns,
    )


def request_aladin_items(url, params):
    try:
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        data = response.json()
    except requests.RequestException as exc:
        status_code = getattr(getattr(exc, "response", None), "status_code", None)
        app.logger.warning(
            "Aladin request failed: %s (status=%s)",
            type(exc).__name__,
            status_code or "network",
        )
        raise
    except ValueError:
        app.logger.warning("Aladin response was not valid JSON")
        raise

    if data.get("errorCode"):
        app.logger.warning("Aladin API rejected request: code=%s", data.get("errorCode"))
        raise requests.HTTPError("Aladin API rejected the request")
    return data.get("item", [])


def is_set_product(item):
    title = str(item.get("title", ""))
    return bool(
        "세트" in title
        or "전권" in title
        or re.search(r"(?:전|총)\s*\d+\s*권", title)
        or re.search(r"(?:^|[\s\[(\-])set(?:$|[\s\])\-:])", title, re.IGNORECASE)
    )


def serialize_aladin_books(items):
    catalog_titles, catalog_isbns = load_catalog_index()
    books = []
    for item in items:
        title = item.get("title", "")
        isbn = item.get("isbn13", item.get("isbn", ""))
        books.append(
            {
                "title": title,
                "author": item.get("author", ""),
                "publisher": item.get("publisher", ""),
                "price": item.get("priceStandard", 0),
                "salePrice": item.get("priceSales", 0),
                "cover": item.get("cover", ""),
                "description": item.get("description", ""),
                "isbn": isbn,
                "link": item.get("link", ""),
                "categoryName": item.get("categoryName", ""),
                "pubDate": item.get("pubDate", ""),
                "isDuplicate": is_catalog_duplicate(
                    title,
                    isbn,
                    catalog_titles,
                    catalog_isbns,
                ),
            }
        )
    return books


def load_persisted_bestsellers():
    ensure_database_ready()
    cache_row = db.session.get(ApiCache, "children_bestsellers")
    if not cache_row:
        return [], None
    try:
        items = json.loads(cache_row.value_json)
    except (TypeError, ValueError):
        return [], cache_row.updated_at
    return items if isinstance(items, list) else [], cache_row.updated_at


def save_persisted_bestsellers(items):
    cache_row = db.session.get(ApiCache, "children_bestsellers")
    if cache_row is None:
        cache_row = ApiCache(key="children_bestsellers", value_json="[]")
        db.session.add(cache_row)
    cache_row.value_json = json.dumps(items, ensure_ascii=False)
    cache_row.updated_at = datetime.now()
    db.session.commit()


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def extract_catalog_rows(rows):
    matrix = [list(row) for row in rows]
    title_col = None
    isbn_col = None
    header_row = None

    for row_index, row in enumerate(matrix[:20]):
        for col_index, raw_value in enumerate(row):
            value = cell_text(raw_value)
            lowered = value.lower()
            if any(token in value for token in ["도서명", "서명", "제목", "자료명"]):
                title_col = col_index
                header_row = row_index
            elif "isbn" in lowered:
                isbn_col = col_index
        if title_col is not None:
            break

    if title_col is None or header_row is None:
        raise ValueError("도서명 또는 서명(자료명) 열을 찾지 못했습니다.")

    catalog_rows = []
    for row in matrix[header_row + 1 :]:
        if title_col >= len(row):
            continue
        title = cell_text(row[title_col])
        if not title:
            continue
        isbn = cell_text(row[isbn_col]) if isbn_col is not None and isbn_col < len(row) else ""
        catalog_rows.append({"title": title, "isbn": isbn})
    return catalog_rows


def query_submissions(grade="", class_num=""):
    ensure_database_ready()
    query = Submission.query.order_by(
        Submission.grade.asc(),
        Submission.class_num.asc(),
        Submission.student_number.asc(),
    )
    if grade:
        query = query.filter(Submission.grade == str(grade))
    if class_num:
        query = query.filter(Submission.class_num == str(class_num))
    return [item.to_dict() for item in query.all()]


def load_export_template_workbook():
    if os.path.exists(EXPORT_TEMPLATE_FILE):
        return openpyxl.load_workbook(EXPORT_TEMPLATE_FILE)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = "2026년 황지중앙초 학생, 학부모 구입 희망도서 목록"
    headers = ["순", "도서명", "출판사", "지은이", "수량", "금액(정가)", "할인금액"]
    for index, header in enumerate(headers, start=1):
        worksheet.cell(row=2, column=index, value=header)
    return workbook


def configure_export_sheet(worksheet):
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.35
    worksheet.page_margins.bottom = 0.35
    worksheet.page_margins.header = 0.2
    worksheet.page_margins.footer = 0.2
    worksheet.print_options.horizontalCentered = True
    worksheet.print_title_rows = "1:2"
    worksheet.freeze_panes = "A3"
    worksheet.column_dimensions["A"].width = 8
    worksheet.column_dimensions["B"].width = 42
    worksheet.column_dimensions["C"].width = 14
    worksheet.column_dimensions["D"].width = 14
    worksheet.column_dimensions["E"].width = 7
    worksheet.column_dimensions["F"].width = 12
    worksheet.column_dimensions["G"].width = 11
    worksheet.row_dimensions[1].height = 42
    worksheet.row_dimensions[2].height = 30
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A1"].font = Font(name="맑은 고딕", size=20)

    header_fill = PatternFill(fill_type="solid", start_color="E8C840", end_color="E8C840")
    center_alignment = Alignment(horizontal="center", vertical="center")
    for col in range(1, 8):
        worksheet.cell(row=2, column=col).fill = header_fill
        worksheet.cell(row=2, column=col).alignment = center_alignment


def fill_export_sheet(worksheet, grade, class_num, books, include_grade_class=True):
    configure_export_sheet(worksheet)
    if include_grade_class and grade and class_num:
        worksheet["A1"] = f"2026년 ( {grade} )학년 ( {class_num} )반 학생, 학부모 구입 희망도서 목록"
    else:
        worksheet["A1"] = "2026년 황지중앙초 학생, 학부모 구입 희망도서 목록"

    for row_num in range(3, 43):
        seq = row_num - 2
        book = books[seq - 1] if seq - 1 < len(books) else {}
        price = int(book.get("price", 0) or 0)
        sale_price = int(book.get("salePrice", 0) or 0)
        if not sale_price and price:
            sale_price = int(price * 0.9)

        worksheet.cell(row=row_num, column=1, value=seq)
        worksheet.cell(row=row_num, column=1).alignment = Alignment(horizontal="center", vertical="center")
        worksheet.cell(row=row_num, column=2, value=book.get("title", ""))
        worksheet.cell(row=row_num, column=3, value=book.get("publisher", ""))
        worksheet.cell(row=row_num, column=4, value=book.get("author", ""))
        worksheet.cell(row=row_num, column=5, value=1 if book else None)
        worksheet.cell(row=row_num, column=5).alignment = Alignment(horizontal="center", vertical="center")
        worksheet.cell(row=row_num, column=6, value=price if book else None)
        worksheet.cell(row=row_num, column=7, value=sale_price if book else 0)
        worksheet.cell(row=row_num, column=6).number_format = "#,##0"
        worksheet.cell(row=row_num, column=7).number_format = "#,##0"

    worksheet.cell(row=43, column=2, value="계")
    worksheet.cell(row=43, column=6, value="=SUM(F3:F42)")
    worksheet.cell(row=43, column=6).number_format = "#,##0"
    worksheet.cell(row=43, column=7, value="=SUM(G3:G42)")
    worksheet.cell(row=43, column=7).number_format = "#,##0"


def build_admin_workbook(submissions):
    workbook = load_export_template_workbook()
    template_sheet = workbook.active
    template_sheet.title = "template"

    groups = {}
    for submission in submissions:
        key = (submission["grade"], submission["classNum"])
        groups.setdefault(key, [])
        groups[key].extend(submission.get("books", []))

    if not submissions:
        worksheet = template_sheet
        worksheet.title = "희망도서"
        fill_export_sheet(worksheet, "", "", [], include_grade_class=False)
    elif len(groups) == 1:
        (grade, class_num), books = next(iter(sorted(groups.items())))
        worksheet = template_sheet
        worksheet.title = f"{grade}학년 {class_num}반"
        fill_export_sheet(worksheet, grade, class_num, books, include_grade_class=True)
    else:
        all_books = []
        for submission in submissions:
            all_books.extend(submission.get("books", []))
        worksheet = template_sheet
        worksheet.title = "황지중앙초"
        fill_export_sheet(worksheet, "", "", all_books, include_grade_class=False)

    if "template" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        workbook.remove(workbook["template"])

    return workbook


def require_admin():
    return session.get("is_admin") is True


@app.route("/")
def index():
    ensure_database_ready()
    return render_template(
        "index.html",
        has_api_key=bool(ALADIN_API_KEY),
        grades=GRADE_OPTIONS,
        school_structure=SCHOOL_STRUCTURE,
    )


@app.route("/api/search")
def search_books():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"books": [], "error": "검색어를 입력해 주세요."})

    if not ALADIN_API_KEY:
        return jsonify({"books": [], "error": "알라딘 API 키가 설정되어 있지 않습니다."})

    try:
        items = request_aladin_items(
            ALADIN_SEARCH_URL,
            {
                "ttbkey": ALADIN_API_KEY,
                "Query": query,
                "QueryType": "Keyword",
                "MaxResults": 20,
                "start": 1,
                "SearchTarget": "Book",
                "output": "js",
                "Version": "20131101",
                "Cover": "Big",
            },
        )
    except requests.RequestException:
        return (
            jsonify(
                {
                    "books": [],
                    "error": "알라딘 도서 검색에 연결하지 못했습니다. 잠시 후 다시 시도해 주세요.",
                }
            ),
            502,
        )
    except ValueError:
        return (
            jsonify(
                {
                    "books": [],
                    "error": "알라딘 도서 검색 응답을 읽지 못했습니다. 잠시 후 다시 시도해 주세요.",
                }
            ),
            502,
        )

    single_books = [item for item in items if not is_set_product(item)]
    return jsonify({"books": serialize_aladin_books(single_books)})


@app.route("/api/bestsellers")
def children_bestsellers():
    if not ALADIN_API_KEY:
        return jsonify({"books": [], "error": "알라딘 API 키가 설정되어 있지 않습니다."})

    now = time.monotonic()
    items = bestseller_cache["items"]
    persisted_items, persisted_at = load_persisted_bestsellers()
    persisted_is_fresh = bool(
        persisted_items
        and persisted_at
        and (datetime.now() - persisted_at).total_seconds()
        < BESTSELLER_CACHE_SECONDS
    )

    if not items and persisted_is_fresh:
        items = persisted_items
        bestseller_cache["items"] = items
        bestseller_cache["expires_at"] = now + BESTSELLER_CACHE_SECONDS

    should_refresh = (
        (not items or now >= bestseller_cache["expires_at"])
        and now >= bestseller_cache["retry_after"]
    )
    if should_refresh:
        try:
            items = request_aladin_items(
                ALADIN_LIST_URL,
                {
                    "ttbkey": ALADIN_API_KEY,
                    "QueryType": "Bestseller",
                    "MaxResults": 50,
                    "start": 1,
                    "SearchTarget": "Book",
                    "CategoryId": ALADIN_CHILDREN_CATEGORY_ID,
                    "output": "js",
                    "Version": "20131101",
                    "Cover": "Big",
                },
            )
            save_persisted_bestsellers(items)
            bestseller_cache["items"] = items
            bestseller_cache["expires_at"] = now + BESTSELLER_CACHE_SECONDS
            bestseller_cache["retry_after"] = 0.0
        except (requests.RequestException, ValueError):
            bestseller_cache["retry_after"] = now + BESTSELLER_RETRY_SECONDS
            if not items:
                items = persisted_items

    if not items:
        return (
            jsonify(
                {
                    "books": [],
                    "error": "어린이 베스트셀러를 불러오지 못했습니다. 잠시 후 다시 시도해 주세요.",
                }
            ),
            502,
        )

    return jsonify({"books": serialize_aladin_books(items)})


@app.route("/api/submit", methods=["POST"])
def submit_books():
    ensure_database_ready()
    data = request.get_json() or {}
    grade = str(data.get("grade", "")).strip()
    class_num = str(data.get("classNum", "")).strip()
    student_number = str(data.get("studentNumber", "")).strip()
    books = data.get("books", [])

    if grade not in GRADE_OPTIONS:
        return jsonify({"success": False, "error": "학년을 다시 선택해 주세요."})
    if class_num not in class_options_for_grade(grade):
        return jsonify({"success": False, "error": "반을 다시 선택해 주세요."})
    if student_number not in student_numbers_for_class(grade, class_num):
        return jsonify({"success": False, "error": "번호를 다시 선택해 주세요."})
    if len(books) != 3:
        return jsonify({"success": False, "error": "희망 도서 3권을 모두 선택해 주세요."})

    if Submission.query.filter_by(
        grade=grade, class_num=class_num, student_number=student_number
    ).first():
        return jsonify(
            {
                "success": False,
                "error": f"{grade}학년 {class_num}반 {student_number}번은 이미 제출했습니다.",
            }
        )

    submission = Submission(
        grade=grade,
        class_num=class_num,
        student_number=student_number,
        student_label=f"{grade}학년 {class_num}반 {student_number}번",
        books_json=json.dumps(
            [
                {
                    "title": book.get("title", ""),
                    "author": book.get("author", ""),
                    "publisher": book.get("publisher", ""),
                    "price": int(book.get("price", 0) or 0),
                    "salePrice": int(book.get("salePrice", 0) or 0),
                    "isbn": book.get("isbn", ""),
                }
                for book in books
            ],
            ensure_ascii=False,
        ),
    )
    db.session.add(submission)
    db.session.commit()

    return jsonify(
        {
            "success": True,
            "message": f"{grade}학년 {class_num}반 {student_number}번 신청이 저장되었습니다.",
        }
    )


@app.route("/api/admin/login", methods=["POST"])
def admin_login():
    data = request.get_json() or {}
    password = str(data.get("password", "")).strip()
    if password != ADMIN_PASSWORD:
        session["is_admin"] = False
        return jsonify({"success": False, "error": "비밀번호가 올바르지 않습니다."}), 401
    session["is_admin"] = True
    return jsonify({"success": True})


@app.route("/api/admin/logout", methods=["POST"])
def admin_logout():
    session["is_admin"] = False
    return jsonify({"success": True})


@app.route("/api/admin/submissions")
def admin_submissions():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "관리자 인증이 필요합니다."}), 401
    grade = request.args.get("grade", "").strip()
    class_num = request.args.get("classNum", "").strip()
    return jsonify({"submissions": query_submissions(grade=grade, class_num=class_num)})


@app.route("/api/admin/submissions", methods=["DELETE"])
def clear_submissions():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "愿由ъ옄 ?몄쬆???꾩슂?⑸땲??"}), 401
    deleted_count = Submission.query.delete()
    db.session.commit()
    return jsonify({"success": True, "deletedCount": deleted_count})


@app.route("/api/admin/export")
def admin_export():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "관리자 인증이 필요합니다."}), 401

    grade = request.args.get("grade", "").strip()
    class_num = request.args.get("classNum", "").strip()
    workbook = build_admin_workbook(query_submissions(grade=grade, class_num=class_num))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    filename_parts = ["희망도서_신청결과"]
    if grade:
        filename_parts.append(f"{grade}학년")
    if class_num:
        filename_parts.append(f"{class_num}반")

    return send_file(
        buffer,
        as_attachment=True,
        download_name="_".join(filename_parts) + ".xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/admin/catalog", methods=["GET"])
def get_catalog():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "관리자 인증이 필요합니다."}), 401
    return jsonify({"count": CatalogBook.query.count()})


@app.route("/api/admin/catalog", methods=["DELETE"])
def clear_catalog():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "관리자 인증이 필요합니다."}), 401
    CatalogBook.query.delete()
    db.session.commit()
    return jsonify({"success": True})


@app.route("/api/admin/upload-catalog", methods=["POST"])
def upload_catalog():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "관리자 인증이 필요합니다."}), 401

    uploaded_file = request.files.get("file")
    if not uploaded_file or not uploaded_file.filename:
        return jsonify({"success": False, "error": "파일을 선택해 주세요."})

    filename = uploaded_file.filename.lower()
    try:
        if filename.endswith(".xlsx"):
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
            worksheet = workbook.active
            catalog_rows = extract_catalog_rows(
                worksheet.iter_rows(values_only=True)
            )
        elif filename.endswith(".xls"):
            workbook = xlrd.open_workbook(file_contents=uploaded_file.read())
            worksheet = workbook.sheet_by_index(0)
            catalog_rows = extract_catalog_rows(
                worksheet.row_values(row_index)
                for row_index in range(worksheet.nrows)
            )
        elif filename.endswith(".csv"):
            wrapper = TextIOWrapper(uploaded_file.stream, encoding="utf-8-sig")
            catalog_rows = extract_catalog_rows(csv.reader(wrapper))
        else:
            return jsonify(
                {
                    "success": False,
                    "error": "xlsx, xls 또는 csv 파일만 업로드할 수 있습니다.",
                }
            )
    except Exception as exc:
        return jsonify({"success": False, "error": f"파일 처리 중 오류가 발생했습니다: {exc}"})

    CatalogBook.query.delete()
    for row in catalog_rows:
        title = row["title"].strip()
        if not title:
            continue
        db.session.add(
            CatalogBook(
                title=title,
                normalized_title=normalize_title(title),
                isbn=normalize_isbn(row.get("isbn", "")) or None,
            )
        )
    db.session.commit()

    return jsonify(
        {
            "success": True,
            "message": f"소장 도서 {len(catalog_rows)}권이 등록되었습니다.",
            "count": len(catalog_rows),
        }
    )


@app.route("/api/admin/bestseller-cache", methods=["POST"])
def update_bestseller_cache():
    ensure_database_ready()
    if not require_admin():
        return jsonify({"error": "관리자 인증이 필요합니다."}), 401

    items = (request.get_json() or {}).get("items", [])
    if not isinstance(items, list) or not 1 <= len(items) <= 50:
        return jsonify({"success": False, "error": "1~50권의 목록이 필요합니다."}), 400

    save_persisted_bestsellers(items)
    bestseller_cache["items"] = items
    bestseller_cache["expires_at"] = time.monotonic() + BESTSELLER_CACHE_SECONDS
    bestseller_cache["retry_after"] = 0.0
    return jsonify({"success": True, "count": len(items)})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)
